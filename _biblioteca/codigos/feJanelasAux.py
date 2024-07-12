# importando bibliotecas
from datetime import datetime
import _biblioteca.codigos.feJanelaPrincipal as feJanelaPrincipal
import numpy as np
import os
import openpyxl
import plotly.graph_objs as go
import plotly.graph_objects as px
from plotly.subplots import make_subplots
import pandas as pd
from PyQt5.QtCore import Qt, QTimer
from PyQt5 import QtGui
from PyQt5.QtGui import QPixmap, QIcon, QFont
from PyQt5.QtWidgets import QHBoxLayout, QWidget, QVBoxLayout, QHeaderView, QTableWidgetItem, QMessageBox, QTableWidget, QDialog, QPushButton, QLabel, QFormLayout, QDialogButtonBox, QComboBox, QComboBox, QLineEdit
from itertools import accumulate
from PyQt5.QtWebEngineWidgets import QWebEngineView
from openpyxl import load_workbook
from PIL import Image

# -----------------------------------------------------------------------------------------------
# classe para criação de janela de gestor
class JanelaSelecionaGestor(QDialog):
    def __init__(self, parent = None):
        super().__init__(parent)
        # propriedades da janela parent
        self.janelaReferencia = parent

        # inicializando gui
        self.f_inicializaGui(parent)

    # -----------------------------------------
    # função para inicializar interface
    def f_inicializaGui(self, parent):
        
        # título da janela
        self.setWindowTitle('LIDERANÇAS')

        # ocultando botão default de ajuda
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        # criação de layout
        layout = QVBoxLayout()
        linha = QHBoxLayout()
        coluna = QVBoxLayout()
        self.entradaGestorDaVez = self.f_geraEntradaListaSuspensaDistribuicoes(coluna, '→ LIDERANÇA', parent.propriedadesGerais)
        self.entradaGestorDaVez.setCurrentText(parent.propriedadesGerais['gestorDaVez'])
        coluna.addWidget(self.entradaGestorDaVez)
        linha.addLayout(coluna)
        
        # pulando para próxima linha
        layout.addLayout(linha)
        linha = QHBoxLayout()
        
        # adicionando espaço forçadamente # FIXME: entender pq precisei forçar com um qlabel, antes não estava dando espaço
        self.saidaTexto = QLabel('')
        layout.addWidget(self.saidaTexto)  

        # > botões
        self.dialogoOpcoes = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.dialogoOpcoes.accepted.connect(self.accept)
        self.dialogoOpcoes.rejected.connect(self.reject)
        self.dialogoOpcoes.button(QDialogButtonBox.Ok).setText('Confirmar')
        self.dialogoOpcoes.button(QDialogButtonBox.Cancel).setText('Cancelar')
        linha.addWidget(self.dialogoOpcoes)
        layout.addLayout(linha)
        # definindo layout
        self.setLayout(layout)

    # -----------------------------------------
    # função para gerar entrada lista suspensa
    def f_geraEntradaListaSuspensaDistribuicoes(self, coluna, texto, parentDadoLido):
        entradaTextoGerada = QLabel(texto)
        coluna.addWidget(entradaTextoGerada)
        entradaListaSuspensaGerada = QComboBox()
        entradaListaSuspensaGerada.addItems(parentDadoLido['gestoesPossiveis'])
        return entradaListaSuspensaGerada

    # -----------------------------------------
    # função para retornar propriedades para a classe de janela principal
    def f_obtemPropriedades(self):
        return self.entradaGestorDaVez.currentText() 

# ----------------------------------------------
# janela que exibe projetos
class JanelaPts(QDialog):
    def __init__(self, indice, parent=None):
        super().__init__(parent)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setStyleSheet('color: #2C4594; font: 15pt') 
        self.indice = indice
        self.gestores = ['MAURO', 'ROGÉRIO', 'MARCOS', 'ANA']
        
        # Carregar dados do arquivo Excel
        # atualizar 1 vez no mês
        if indice == 0:
            arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos de Trabalho/projetosMauro.xlsx'
        elif indice == 1:
            arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos de Trabalho/projetosRogerio.xlsx'
        elif indice == 2:
            arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos de Trabalho/projetosMarcos.xlsx'
        else:
            arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos de Trabalho/projetosAna.xlsx'
            
        dados = self.f_carregarExcel(arquivo)
        # Exibir dados em um QTableWidget
        self.f_exibirTabela(dados)
        
        self.logoVibracon = None

        # chamando scroll 
        self.scrollTimer = QTimer(self)
        self.scrollTimer.timeout.connect(self.f_scrollDown)
        self.scrollTimer.start(15000)
        self.scrollCont = 1
    
    # ---------------------------------
    # função que carrega os dados do arquivo excel escrito do project
    def f_carregarExcel(self, arquivo):
        wb = openpyxl.load_workbook(arquivo)
        planilha = wb.active
        dados = []
        for linha in planilha.iter_rows(values_only=True):
            dados.append(linha)
        return dados
    
    # -------------------------------------
    # função que exibe tabela
    def f_exibirTabela(self, dados): 
        data = datetime.now()
        dataParaExibir = data.strftime("%Y-%m-%d")
        # > fe
        # título da janela
        self.setWindowTitle(f'VIBRACON VibraPlan (v.{dataParaExibir})')
        self.setGeometry(0, 0, 2500, 2500)
        # ícone
        self.setWindowIcon(QIcon('_biblioteca/arte/logos/logoVibracon1.png'))
                
        # bg
        self.bg = QLabel(self)
        self.bg.setGeometry(0, 0, 2500, 2500)
        pixmap = QPixmap('_biblioteca/arte/bg/bgUm.png')
        self.bg.setPixmap(pixmap)
        self.bg.setScaledContents(True)
        
        # criação de layout
        layout = QVBoxLayout() # vertical
        layout.addSpacing(self.height() * 0.02)

        linha = QHBoxLayout() # horizontal

        self.labelGestor = QLabel(self)
        self.labelGestor.setStyleSheet("font-size: 20pt;")
        self.labelGestor.setText(f'COORDENADOR(A): {self.gestores[self.indice]}')
        linha.addWidget(self.labelGestor)
        
        linha.addStretch()
        # logo vibracon 
        caminho = '_biblioteca/arte/logos/logoPrograma.png'
        self.logoVibracon = QLabel()
        pixmap = QPixmap(caminho)
        self.logoVibracon.setPixmap(pixmap)
        self.logoVibracon.setMaximumHeight(80)
        self.logoVibracon.setMaximumWidth(1000)
        linha.addWidget(self.logoVibracon)
        linha.setAlignment(Qt.AlignRight)

        # pulando para próxima linha
        layout.addLayout(linha)
        linha = QHBoxLayout()
        
        # espaçamento vertical
        layout.addSpacing(int(self.height() * 0.02))

        # canvas para tabela
        self.quadroTarefas = QTableWidget()
        layout.addWidget(self.quadroTarefas)

        # pulando para próxima linha
        # espaçamento vertical
        layout.addSpacing(int(self.height() * 0.01))

        # criando legenda para status
        linha = QHBoxLayout()
        linha.addStretch()
        
        legenda = {'ENTREGUE': '#5D9145', 'ATRASADO': '#CF0E0E', 'RISCO': '#FFCC33'}
        # Adicionando QLabel para cada entrada na legenda
        for texto, cor in legenda.items():
            # Criando o ponto com a cor do status
            ponto = QLabel()
            ponto.setStyleSheet(f"background-color: {cor}; border-radius: 5%;")
            linha.addWidget(ponto)

            # Criando QLabel com o texto
            textoLabel = QLabel(f"{texto}   ")
            textoLabel.setStyleSheet("font-size: 12pt;")
            textoLabel.setFont(QFont('Gill Sans MT'))
            linha.addWidget(textoLabel)
            
        layout.addLayout(linha)
        layout.addSpacing(int(self.height() * 0.005))

        self.showMaximized()
        self.setLayout(layout)
        # limpando tabela
        self.quadroTarefas.clearContents()
        self.quadroTarefas.setRowCount(0)
        self.quadroTarefas.setColumnCount(0)

        # estilizando colunas
        self.quadroTarefas.setColumnCount(4)
        self.quadroTarefas.setHorizontalHeaderLabels(['PROJETO', 'PRAZO', ' % ', 'STATUS'])
        self.quadroTarefas.setColumnWidth(0, 940)
        self.quadroTarefas.setColumnWidth(1, 492)
        self.quadroTarefas.setColumnWidth(2, 200)
        self.quadroTarefas.setColumnWidth(3, 197)
        self.quadroTarefas.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.quadroTarefas.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.quadroTarefas.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.quadroTarefas.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        header = self.quadroTarefas.horizontalHeader()
        font = header.font()
        font.setPointSize(20)
        header.setFont(font)

        # Preencher a tabela com os dados
        for row in dados:
            linhas = self.quadroTarefas.rowCount()
            self.quadroTarefas.insertRow(linhas)
            for coluna, valor in enumerate(row):
                if coluna == 2:
                    # Concatena o valor com '%' 
                    valor = f"{valor}%"
                item = QTableWidgetItem(str(valor))
                if coluna in (1, 2):  # Verifica se a coluna é 1 ou 2
                    item.setTextAlignment(Qt.AlignCenter)  # Alinha o texto ao centro
                self.quadroTarefas.setRowHeight(linhas, 50)
                self.quadroTarefas.setItem(linhas, coluna, item)


        for linhaDaVez in range(self.quadroTarefas.rowCount()):
            celulaDaVez = self.quadroTarefas.item(linhaDaVez, 0)
            textoCelulas = celulaDaVez.text()
            if textoCelulas.startswith(' '):
                pass
            else:
                celulaDaVez.setBackground(QtGui.QColor('#2C4594'))
                celulaDaVez.setForeground(QtGui.QColor('white'))
                self.quadroTarefas.setItem(linhaDaVez, 1, QTableWidgetItem(''))
                self.quadroTarefas.setItem(linhaDaVez, 2, QTableWidgetItem(''))
            celulaStatus = self.quadroTarefas.item(linhaDaVez, 3)
            textoCelula = ''
            textoCelula = celulaStatus.text()
            if textoCelula is not None:
                if textoCelula == 'atrasado':
                    celulaStatus.setBackground(QtGui.QColor('#CF0E0E'))
                    celulaStatus.setForeground(QtGui.QColor('#CF0E0E'))
                elif textoCelula == 'RISCO':
                    celulaStatus.setBackground(QtGui.QColor('#FFCC33'))
                    celulaStatus.setForeground(QtGui.QColor('#FFCC33'))
                elif textoCelula == 'no prazo':
                    celulaStatus.setBackground(QtGui.QColor('#5D9145'))
                    celulaStatus.setForeground(QtGui.QColor('#5D9145'))
                else:
                    self.quadroTarefas.setItem(linhaDaVez, 3, QTableWidgetItem(''))

    # ----------------------------------------
    # função do scroll automático 
    def f_scrollDown(self):
            # definindo scroll automático
            scroll = self.quadroTarefas.verticalScrollBar()
            valorMax = scroll.maximum()
            valorAtual = scroll.value()
            self.scrollTimer.setInterval(2000)
            novoValor = min(valorAtual + self.scrollCont, valorMax)
            scroll.setValue(novoValor)

            # condição de parada do scroll
            if novoValor == valorMax:
                self.scrollTimer.timeout.disconnect(self.f_scrollDown)
                self.scrollTimer.stop() 
                janelaPrincipal = feJanelaPrincipal.JanelaPrincipal(self.indice)
                janelaPrincipal.showMaximized()
                self.close()

# ----------------------------------
# classe do print do trelo ou outro print (powerbi, logo vibracon)
class JanelaImagem(QDialog):
    def __init__(self, indice, parent=None):
        super().__init__(parent)
        self.indice = indice
        # verificar se é o mauro para exibir o print do trelo
        if self.indice != 0:
            # Criar um timer para fechar a janela e abrir JanelaPts após x segundos
            self.timer = QTimer(self)
            self.timer.setSingleShot(True)  # Executa apenas uma vez
            self.timer.timeout.connect(self.f_abrirJanelaPts)
            self.timer.start(500) 
        else:
            # Criar um timer para fechar a janela e abrir JanelaPts após x segundos
            self.timer = QTimer(self)
            self.timer.setSingleShot(True)  # Executa apenas uma vez
            self.timer.timeout.connect(self.f_abrirJanelaPts)
            self.timer.start(20000) 
        
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setWindowTitle('Vibraplan')
        self.setGeometry(100, 100, 1760, 990)
        
        # Configurar layout
        layout = QVBoxLayout()
        
        # Carregar imagem
        imagem = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Documentos/printTrelo.jpg'
        self.imagemLabel = QLabel(self)
        pixmap = QPixmap(imagem)
        self.imagemLabel.setPixmap(pixmap)
        self.imagemLabel.setScaledContents(True)
        
        # Adicionar imagem ao layout
        layout.addWidget(self.imagemLabel)
        
        # Configurar layout na janela
        self.setLayout(layout)
        self.showMaximized()

    # ---------------------------------------
    # Função para fechar a janela de imagem e abrir a JanelaPts
    def f_abrirJanelaPts(self):
        # chama a janela de projetos
        if self.indice == 3:
            janelaPts = JanelaPts(0)
        else:
            janelaPts = JanelaPts(self.indice+1)
        janelaPts.showMaximized()
        janelaPts.exec_()
        self.close()

# ----------------------------------------------
# janela que exibe gráficos
class JanelaGraficos(QDialog):
    def __init__(self, graficoAderencia, graficoAderenciaGeral, graficoDesvio, graficoCoordenadores, graficoProjetos, gestorDaVez, indice, parent=None):
        super().__init__(parent)
        self.indiceGeral = indice
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setWindowTitle('Vibraplan')
        self.setGeometry(100, 100, 1760, 990)
        # ícone
        self.setWindowIcon(QIcon('_biblioteca/arte/logos/logoVibracon1.png'))
        self.gestorDaVez = gestorDaVez
        # bg
        self.bg = QLabel(self)
        self.bg.setGeometry(0, 0, 2500, 2500)
        pixmap = QPixmap('_biblioteca/arte/bg/bgUm.png')
        self.bg.setPixmap(pixmap)
        self.bg.setScaledContents(True)

        # Salvando os gráficos HTML
        self.graficoAderenciaHtml = graficoAderencia
        self.graficoAderenciaGeralHtml = graficoAderenciaGeral
        self.graficoDesvioHtml = graficoDesvio
        self.graficoCoordenadoresHtml = graficoCoordenadores
        self.graficoProjetosHtml = graficoProjetos

        # Criar uma visualização do WebEngine para mostrar o gráfico
        self.webview = QWebEngineView()
        # Criar um layout para a janela
        layout = QVBoxLayout()
        layout.addWidget(self.webview)

        # Definir o layout para a janela
        self.setLayout(layout)

        self.showMaximized()
        
        # Inicializar o índice do gráfico atual
        self.indiceGrafico = 0
        # Iniciar o timer para alternar entre os gráficos a cada 30 segundos
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.f_exibirProximoGrafico)
        self.timer.start(10000)

        # Exibir o primeiro gráfico
        self.f_exibirProximoGrafico()

    # ---------------------------------
    # função que troca os gráficos
    def f_exibirProximoGrafico(self):
        try:
            # Lista com os gráficos HTML
            graficos = [self.graficoAderenciaHtml, self.graficoDesvioHtml]
            graficosFinais = [self.graficoAderenciaGeralHtml, self.graficoProjetosHtml, self.graficoCoordenadoresHtml]

            # Verificar se todos os gráficos foram exibidos
            if self.indiceGrafico >= len(graficos):
                janelamagem = JanelaImagem(self.indiceGeral)
                janelamagem.exec_()
                self.close()
            # Exibir o próximo gráfico na lista
            self.webview.setHtml(graficos[self.indiceGrafico])

            # Atualizando indice
            self.indiceGrafico += 1
        except Exception: pass
# ----------------------------------------------
# função para geração dos gráficos de aderências
def f_plotaAderencias(self, datasCompletas, contadorDesvio):
    indiceTabela = 0
    for indice, gestao in enumerate(self.propriedadesGerais['gestoesPossiveis']):
        if gestao == self.propriedadesGerais['gestorDaVez']:
            indiceTabela = indice
            break 
    # obtendo número de dias por semana       
    diasSemanaVez = 0
    diferencaEmDias = 0
    diasPorSemana = []
    totalAderencias = 0
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    # Obtendo o nome do mês atual em inglês
    # achando o mês atual
    for data in self.tabelaLida[indiceTabela]['dados']['datas']:
        if not pd.isna(data):
            if isinstance(data, pd.Timestamp):
                # Extrair o mês da data válida
                mes = data.month
                break
            elif isinstance(data, str):
                # Converter a string para data e extrair o mês
                dataConvertida = datetime.strptime(data, '%d-%b-%y')
                mes = dataConvertida.month
                break
        # Convertendo o nome do mês para português
    mesEscrito = meses[mes-1].upper()
    for indiceData in range(len(datasCompletas)-1):
        diferencaEmDias = abs((datetime.strptime(datasCompletas[indiceData+1], '%d-%b') - datetime.strptime(datasCompletas[indiceData], '%d-%b')).days)
        if diferencaEmDias <= 1:
            diasSemanaVez += 1
            pass
        else:
            diasPorSemana.append(diasSemanaVez+1)           
            diasSemanaVez = 0
    diasPorSemana.append(diasSemanaVez+1)
    diasPorSemanaTotal = [i+2 for i in diasPorSemana]

    # verificando quantidade de ok e de entregue por dia
    contaTarefas = 0
    listaTotalTarefas = []
    contaEntregue = 0
    totalEntregue = []

    for colunaDaVez in range(self.quadroTarefas.columnCount() - 7):
        for linhaDaVez in range(self.quadroTarefas.rowCount()):
            try:
                if self.quadroTarefas.item(linhaDaVez, colunaDaVez + 1).text() != '' and self.quadroTarefas.item(linhaDaVez, colunaDaVez + 1).text() != 'SD':
                    if self.tabelaLida[indiceTabela]['dados'].iloc[linhaDaVez, -6] == 'Aprovação do cliente':
                        pass
                    elif self.tabelaLida[indiceTabela]['dados'].iloc[linhaDaVez, -6] == 'Solicitação de exclusão pelo cliente':
                        pass
                    elif self.tabelaLida[indiceTabela]['dados'].iloc[linhaDaVez, -6] == 'Falta de informação':
                        pass
                    elif self.tabelaLida[indiceTabela]['dados'].iloc[linhaDaVez, 1] == 'INATIVO':
                        pass 
                    else:
                        contaTarefas += 1
                        # verificando quantidade de entregue
                    
                        if self.tabelaLida[indiceTabela]['dados']['status'][linhaDaVez] == 'ENTREGUE' :
                            contaEntregue += 1
            except: pass
        listaTotalTarefas.append(contaTarefas)
        contaTarefas = 0
        totalEntregue.append(contaEntregue)
        totalAderencias += contaEntregue
        contaEntregue = 0
    # verificando quantidade de ok e de entregue por semana
    auxOk = 0
    indiceOk = 0
    okSemanal = []
    auxEntregue = 0
    indiceEntregue = 0
    entregueSemanal = []
    for diasSemanaVez in diasPorSemana:
        while diasSemanaVez > 0:
            if indiceOk == len(listaTotalTarefas):
                break
            auxOk += listaTotalTarefas[indiceOk]
            indiceOk += 1
            auxEntregue += totalEntregue[indiceEntregue]
            indiceEntregue += 1
            diasSemanaVez -= 1
        okSemanal.append(auxOk)
        auxOk = 0
        entregueSemanal.append(auxEntregue)
        auxEntregue = 0

    # cálculos
    # Convertendo cada valor do dicionário para um inteiro
    for chave, valor in self.tarefasAcumuladas.items():
        if not isinstance(valor, int):
            self.tarefasAcumuladas[chave] = int(valor[0])
    while len(entregueSemanal) < 5:
        entregueSemanal.append(0)
    while len(okSemanal) < 5:
        okSemanal.append(0)
    efprazoCorreto = [0, 0, 0, 0, 0]
    okSemanal = np.array(okSemanal)
    entregueSemanal = np.array(entregueSemanal)
    # aderência semanal
    aderenciaSemanal = np.round(100 * entregueSemanal/okSemanal, 1)
    # aderência acumulada
    okAcumulado = np.array(list(accumulate(okSemanal)))

    for valor in self.tarefasAcumuladas.values():
        soma = 0
        i = 0
        for i, numDias in enumerate(diasPorSemanaTotal):
            soma += numDias
            if soma >= int(valor):
                efprazoCorreto[i] += 1
                break
    while len(efprazoCorreto) < 5:
        efprazoCorreto.append(0)
    efprazoCorreto = np.array(efprazoCorreto)
    entreguesTotal = entregueSemanal + efprazoCorreto 
    entregueComEfprazoAcumulado = np.array(list(accumulate(entreguesTotal)))
    aderenciaAcumulada = np.round(100 * entregueComEfprazoAcumulado/okAcumulado, 1)
    # escrevendo as aderências num arquivo .txt
    caminho = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Aderencias'
    arquivo = os.path.join(caminho, f"Aderencias{self.propriedadesGerais['gestorDaVez']}.txt")

    # Lista para armazenar as aderências semanais e acumuladas
    aderenciasSemanais = []
    aderenciasAcumuladas = []

    # Navegue até a pasta
    os.chdir(caminho)

    # Obtenha a lista de arquivos na pasta
    arquivos = os.listdir()
    arquivosTxt = [arquivo for arquivo in arquivos if arquivo.endswith('.txt')]

    # Obter o nome do gestor atual
    gestorDaVez = self.propriedadesGerais['gestorDaVez']

    # Lista para armazenar as aderências semanais e acumuladas
    aderenciasSemanais = []
    aderenciasAcumuladas = []
    semanalDaVez = []
    acumuladaDaVez = []

    try:
        # Navegue até a pasta
        os.chdir(caminho)
        # Obtenha a lista de arquivos na pasta
        arquivos = os.listdir()
        arquivosTxt = [arquivo for arquivo in arquivos if arquivo.endswith('.txt')]
        for arquivoTxt in arquivosTxt:
            with open(arquivoTxt, 'r') as arquivo:
                # Leia os dados do arquivo e armazene-os na lista
                dados = arquivo.readlines()
                # Supondo que as aderências semanais estão nas primeiras linhas e as acumuladas nas linhas seguintes
                aderenciasSemanais.append([float(aderencia.strip()) for aderencia in dados[:len(dados)//2]])
                aderenciasAcumuladas.append([float(aderencia.strip()) for aderencia in dados[len(dados)//2:]])
                while len(aderenciasSemanais[-1]) < 5:
                    aderenciasSemanais[-1].append(0.0)
                while len(aderenciasAcumuladas[-1]) < 5:
                    aderenciasAcumuladas[-1].append(0.0)
    except Exception as erro: QMessageBox.critical(self, 'AVISO', f'Erro ao adicionar.\n\n {str(erro)}')
    
    # Exibindo as aderências semanais e acumuladas
    totalSemanal = np.mean(aderenciasSemanais, axis=0)
    totalAcumulada = np.mean(aderenciasAcumuladas, axis=0)

    # gerando lista de semanas
    nomesSemana = []
    for nSemana in range(5):
        nomesSemana.append('Semana ' + str(nSemana+1))
    indiceAderencia = 0
    if indiceTabela == 0:
        indiceAderencia = 3
    elif indiceTabela == 1:
        indiceAderencia = 4
    elif indiceTabela == 2:
        indiceAderencia = 2
    elif indiceTabela == 3:
        indiceAderencia = 0
    semanalDaVez = aderenciasSemanais[indiceAderencia]
    acumuladaDaVez = aderenciasAcumuladas[indiceAderencia]
    semanalDaVez = np.array(semanalDaVez)
    acumuladaDaVez = np.array(acumuladaDaVez)

    # plotando
    # curva de meta
    curvaMeta = go.Scatter(
        x = [nomesSemana[0], nomesSemana[-1]],
        y = [80, 80],
        name = 'Meta',
        marker = dict(color = 'rgba(93, 145, 69, .5)', size = 0),
        mode = 'lines+markers',
    )

    # criando figura com subplots lado a lado
    graficoAderencia = make_subplots(
        rows=1, cols=2,  # Uma linha, duas colunas
        subplot_titles=['Semanal [%]', 'Acumulada [%]'],
    )
    textoAderenciaSemanal = [f'{valor} %' for valor in aderenciasSemanais[indiceAderencia]]
    # Adiciona o gráfico semanal ao primeiro subplot
    aderenciaSemanal = go.Bar(
        x=nomesSemana,
        y=aderenciasSemanais[indiceAderencia],
        name='Semanal',
        marker=dict(color='rgba(44, 69, 148, 1)'),
        text=textoAderenciaSemanal
    )
    graficoAderencia.add_trace(aderenciaSemanal, row=1, col=1)  # Adiciona ao primeiro subplot
    textoAderenciaAcumulada = [f'{valor} %' for valor in aderenciasAcumuladas[indiceAderencia]]
    # Adiciona o gráfico acumulado ao segundo subplot
    aderenciaAcumulada = go.Bar(
        x=nomesSemana,
        y=aderenciasAcumuladas[indiceAderencia], 
        name='Acumulada',
        marker=dict(color='rgba(44, 69, 148, 1)'),
        text=textoAderenciaAcumulada,
    )
    graficoAderencia.add_trace(aderenciaAcumulada, row=1, col=2)  # Adiciona ao segundo subplot

    # Adiciona a curva de meta aos dois subplots
    curvaMeta = go.Scatter(
        x=[nomesSemana[0], nomesSemana[-1]],
        y=[85, 85],
        name='Meta',
        marker=dict(color='rgba(93, 145, 69, .5)', size=0),
        mode='lines+markers',
    )
    graficoAderencia.add_trace(curvaMeta, row=1, col=1)
    graficoAderencia.add_trace(curvaMeta, row=1, col=2)
    
    # Ajusta o layout do gráfico
    graficoAderencia.update_layout(
        barmode='group',  # Define o modo de agrupamento para barras
        showlegend=False,
        title_text=f'ADERÊNCIAS {mesEscrito}/24: ' + self.propriedadesGerais['gestorDaVez'].upper(),
        width=1950,
        height=950,
        title={
            'font': dict(
                size=24,
                family= 'Gill Sans MT'
            )
        }
    )


    #cálculos para gráfico de desvio 
    totalTarefas = 0 
    totalTarefas = np.sum(okSemanal)

    # Obtendo o dicionário de contadores de desvios
    totalNaoEntregue = sum(contadorDesvio.values())
    porcentagensDesvios = []
    for valor in contadorDesvio.values():
        if valor != 0:
            porcentagensDesvios.append(valor/totalNaoEntregue * 100)
    motivos = [motivo for motivo, valor in contadorDesvio.items() if valor > 0]
    # verificação dos desvios que não são contabilizados
    if 'Aprovação do cliente' in motivos:
        motivos.remove('Aprovação do cliente')
    if 'Solicitação de exclusão pelo cliente' in motivos:
        motivos.remove('Solicitação de exclusão pelo cliente')
    if 'Falta de informação' in motivos:
        motivos.remove('Falta de informação')
    
    # Ordenar os motivos e valores em ordem decrescente
    motivosOrdenados = sorted(motivos, key=lambda x: contadorDesvio[x], reverse=True)
    valoresOrdenados = [contadorDesvio[motivo] for motivo in motivosOrdenados]
    porcentagensAcumuladas = np.cumsum(valoresOrdenados) / totalTarefas * 100
    total = sum(valoresOrdenados)
    porcentagensAcumuladas = [sum(valoresOrdenados[:i+1]) / total * 100 for i in range(len(valoresOrdenados))]
    porcentagensIndiv = [valor / total * 100 for valor in valoresOrdenados] # Calculando porcentagens individuais para cada barra
    labels = [f'{porcentagens:.1f}%' for porcentagens in porcentagensIndiv]

    widthBarra = 0.3
    if len(motivosOrdenados) in [2, 3]:
        widthBarra = 0.35
    elif len(motivosOrdenados) > 3:
        widthBarra = 0.5

    # Criando o gráfico de barras
    graficoDesvio = go.Figure()
    graficoDesvio.add_trace(go.Bar(
        x=motivosOrdenados,
        y=valoresOrdenados,
        name='Quantidade',
        text=labels,
        textposition='auto',
        marker=dict(color='rgb(44, 69, 148, 1)'),
        width=widthBarra
    ))
    try:
        valorMaximo = max(valoresOrdenados)
        graficoDesvio.update_yaxes(range=[0, valorMaximo], dtick=1)
    except ValueError:  
        graficoDesvio.update_yaxes(range=[0, 1], dtick=1)

    # Adicionando o gráfico de linha para as porcentagens acumuladas
    graficoDesvio.add_trace(go.Scatter(
        x=motivosOrdenados,
        y=porcentagensAcumuladas,
        mode='lines+markers',
        name='Porcentagem Acumulada %',
        yaxis='y2',
        line=dict(color='rgb(255, 0, 0)'),
        marker=dict(color='rgb(255, 0, 0)')
    ))
    # Layout do gráfico
    graficoDesvio.update_layout(
        legend=dict(
            font=dict(size=15)),
        barmode='group',
        title_text=f"RELAÇÃO DE DESVIOS {mesEscrito}/24: {self.propriedadesGerais['gestorDaVez']}",
        xaxis=dict(title='Desvios', tickfont=dict(size=15), title_font=dict(size=20)),
        yaxis=dict(title='Quantidade'),
        yaxis2=dict(
            title='Porcentagem Acumulada %',
            overlaying='y',
            side='right'
        ),
        title_font_size=24,
        width=1850,
        height=950
    )

    # gráfico da engenharia total
    # Gerando lista de semanas
    nomesSemana = ['Semana ' + str(i + 1) for i in range(len(totalSemanal))]
    # Criando o gráfico
    graficoEngenharia = make_subplots(rows=1, cols=2, subplot_titles=['Semanal [%]', 'Acumulada [%]'])
    # Adicionando as barras ao gráfico com os valores de totalSemanal
    graficoEngenharia.add_trace(
        go.Bar(x=nomesSemana, y=totalSemanal, name='Semanal [%]', 
        marker=dict(color='rgba(44, 69, 148, 1)'), text=[f"{round(valor, 1)} %\n" for valor in totalSemanal], textposition='auto'), 
        row=1, col=1)

    # Adicionando as barras ao gráfico com os valores de totalAcumulada
    graficoEngenharia.add_trace(
        go.Bar(x=nomesSemana, y=totalAcumulada, name='Acumulada [%]', 
        marker=dict(color='rgba(44, 69, 148, 1)'), text=[f"{round(valor, 1)} %\n" for valor in totalAcumulada], textposition='auto'), 
        row=1, col=2)

    # Ajustando o layout do gráfico
    graficoEngenharia.update_layout(
        barmode='group', 
        showlegend=False,
        title_text='ADERÊNCIAS TOTAIS DA ENGENHARIA',
        width=1870,
        height=970, 
        title_font_size=24,
    )
    graficoEngenharia.update_yaxes(range=[0, 90], dtick=10)

    return graficoAderencia, graficoEngenharia, graficoDesvio
    
# -----------------------------------------
# função para plotar o gráfico pizza
def f_plotaTarefas(self): 
    # separando os dados da planilha
    planilha = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Documentos/informativo.xlsx'  
    df = pd.read_excel(planilha, usecols=[1])
    valoresTarefas = df.iloc[0:4, 0].tolist()
    valoresEng = df.iloc[13:16, 0].tolist()
    valoresEngTotais = [f'NO PRAZO: {int (valoresEng[0])}', f'ATENÇÃO: {int (valoresEng[1])}', f'ATRASADO: {int (valoresEng[2])}']

    # nomeando líderes
    nomesLideres = [f'MAURO: {int (valoresTarefas[0])}', f'ROGÉRIO: {int (valoresTarefas[1])}', 
                    f'MARCOS: {int (valoresTarefas[2])}', f'ANA: {int (valoresTarefas[3])}']
    status = ['NO PRAZO', 'ATENÇÃO', 'ATRASADO']
    coresTarefas = ['#3169A0', '#2C4594', '#368CAB', '#347BA6']
    cores = ['#5D9145', '#EE964B', '#DF6158']

    # criando os gráficos principais
    graficoProjetos = make_subplots(rows=1, cols=2, specs=[[{'type':'domain'}, {'type':'domain'}]],
                                    subplot_titles=['RELAÇÃO DE PROJETOS POR COORDENADOR <br> ', 'STATUS DOS PROJETOS DA ENGENHARIA <br> '])
    graficoProjetos.add_trace(go.Pie(labels=nomesLideres, values=valoresTarefas, textposition='inside', scalegroup='one', textinfo='label+percent', 
                                     insidetextfont=dict(size=18, color='white'), titlefont=dict(size=30),
                                     showlegend=False, marker=dict(colors=coresTarefas)), row=1, col=1)
    graficoProjetos.add_trace(go.Pie(labels=valoresEngTotais, values=valoresEng, textposition='inside', scalegroup='one', textinfo='label+percent',
                                     insidetextfont=dict(size=18, color='white'), showlegend=False,
                                     marker=dict(colors=cores)), row=1, col=2)

    graficoProjetos.update_annotations(font=dict(family='Gill Sans MT', size=22, color='#2c4594'))
    graficoProjetos.update_layout(margin=dict(t=200, b=90, l=0, r=70),
        height=900,
        title={
            'text': 'RESUMO EXECUTIVO',
            'y':0.95,
            'x':0.05,
            'font': {
            'family': 'Gill Sans MT',
            'color': '#2c4594',
            'size': 28
            }}
    )
    # definindo imagem de bg
    bg = Image.open('W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_biblioteca/arte/bg/bgUm.png') # tive que usar a PIL para converter a foto
    graficoProjetos.add_layout_image(
        dict(
            source=bg,
            xref="paper",
            yref="paper",
            x=0,
            y=1,
            sizex=2,
            sizey=2,
            sizing="stretch",
            opacity=0.5,
            layer="below"
        )
    )
    # Setando templates
    graficoProjetos.update_layout(template="plotly_white")
    

    # criando os gráficos dos coordenadores
    mauro = df.iloc[19:22, 0].tolist()
    rogerio = df.iloc[25:28, 0].tolist()
    marcos = df.iloc[31:34, 0].tolist()
    ana = df.iloc[37:40, 0].tolist()

    # extraindo dados para hoover do mauro 
    dadosMauro = pd.read_excel(planilha, sheet_name='mauro')
    dataBaseMauro = dadosMauro.iloc[:, 1]
    dataPlanejadaMauro = dadosMauro.iloc[:, 3]

    # inicializando as listas
    atrasoMauro = []
    riscoMauro = []
    okMauro = []

    # Iterar sobre as datas e compará-las
    for i, (dataBase, dataPlanejada) in enumerate(zip(dataBaseMauro, dataPlanejadaMauro)):
        diferencaDias = (dataPlanejada - dataBase).days
        pt = dadosMauro.iloc[i, 0]
        if diferencaDias > 14:
            atrasoMauro.append(pt)
        elif diferencaDias > 1:
            riscoMauro.append(pt)
        else:
            okMauro.append(pt)

    contagensMauro = [len(okMauro), len(riscoMauro), len(atrasoMauro)]
    textoMauro = [f'{s}: {c}' for s, c in zip(status, contagensMauro)]    
    hoverMauro = [okMauro, riscoMauro, atrasoMauro]
    hoverTemplateMauro = [
    '<br>'.join(str(item) for item in sublist) for sublist in hoverMauro
    ]

    i=0
    dataBase=0
    dataPlanejada=0
    diferencaDias=0
    pt= ''

    # extraindo dados para hoover do Rogério 
    dadosRogerio = pd.read_excel(planilha, sheet_name='rogério')
    dataBaseRogerio = dadosRogerio.iloc[:, 1]
    dataPlanejadaRogerio = dadosRogerio.iloc[:, 3]

    # inicializando as listas
    atrasoRogerio = []
    riscoRogerio = []
    okRogerio = []

    # Iterar sobre as datas e compará-las
    for i, (dataBase, dataPlanejada) in enumerate(zip(dataBaseRogerio, dataPlanejadaRogerio)):
        diferencaDias = (dataPlanejada - dataBase).days
        pt = dadosRogerio.iloc[i, 0]
        if diferencaDias > 10:
            atrasoRogerio.append(pt)
        elif diferencaDias > 1:
            riscoRogerio.append(pt)
        else:
            okRogerio.append(pt)

    contagensRogerio = [len(okRogerio), len(riscoRogerio), len(atrasoRogerio)]
    textoRogerio = [f'{s}: {c}' for s, c in zip(status, contagensRogerio)]
    hoverRogerio = [okRogerio, riscoRogerio, atrasoRogerio]
    hoverTemplateRogerio = [
    '<br>'.join(str(item) for item in sublist) for sublist in hoverRogerio
    ]

    i=0
    dataBase=0
    dataPlanejada=0
    diferencaDias=0
    pt= ''

    # extraindo dados para hoover do Marcos
    dadosMarcos = pd.read_excel(planilha, sheet_name='marcos')
    dataBaseMarcos = dadosMarcos.iloc[:, 1]
    dataPlanejadaMarcos = dadosMarcos.iloc[:, 3]

    # inicializando as listas
    atrasoMarcos = []
    riscoMarcos = []
    okMarcos = []

    # Iterar sobre as datas e compará-las
    for i, (dataBase, dataPlanejada) in enumerate(zip(dataBaseMarcos, dataPlanejadaMarcos)):
        diferencaDias = (dataPlanejada - dataBase).days
        pt = dadosMarcos.iloc[i, 0]
        if diferencaDias > 10:
            atrasoMarcos.append(pt)
        elif diferencaDias > 1:
            riscoMarcos.append(pt)
        else:
            okMarcos.append(pt)

    contagensMarcos = [len(okMarcos), len(riscoMarcos), len(atrasoMarcos)]
    textoMarcos = [f'{s}: {c}' for s, c in zip(status, contagensMarcos)]
    hoverMarcos = [okMarcos, riscoMarcos, atrasoMarcos]
    hoverTemplateMarcos = [
    '<br>'.join(str(item) for item in sublist) for sublist in hoverMarcos
    ]

    i=0
    dataBase=0
    dataPlanejada=0
    pt= ''

    # extraindo dados para hoover da ana
    dadosAna = pd.read_excel(planilha, sheet_name='ana')
    dataBaseAna = dadosAna.iloc[:, 1]
    dataPlanejadaAna = dadosAna.iloc[:, 3]

    # inicializando as listas
    atrasoAna = []
    riscoAna = []
    okAna = []

    # Iterar sobre as datas e compará-las
    for i, (dataBase, dataPlanejada) in enumerate(zip(dataBaseAna, dataPlanejadaAna)):
        diferencaDias = (dataPlanejada - dataBase).days
        pt = dadosAna.iloc[i, 0]
        if diferencaDias > 14:
            atrasoAna.append(pt)
        elif diferencaDias > 1:
            riscoAna.append(pt)
        else:
            okAna.append(pt)

    contagensAna = [len(okAna), len(riscoAna), len(atrasoAna)]
    textoAna = [f'{s}: {c}' for s, c in zip(status, contagensAna)]
    hoverAna = [okAna, riscoAna, atrasoAna]
    hoverTemplateAna = [
    '<br>'.join(str(item) for item in sublist) for sublist in hoverAna
    ]

    # Criando subplots
    graficoCoordenadores = make_subplots(rows=1, cols=4, subplot_titles=['MAURO', 'ROGÉRIO', 'MARCOS', 'ANA'], column_widths=[1, 1, 1, 1],
                                        specs=[[{"type": "domain"}, {"type": "domain"}, {"type": "domain"}, {"type": "domain"}]])

    # Adicionando os traces
    graficoCoordenadores.add_trace(go.Pie(labels=status, values=mauro, hovertemplate=hoverTemplateMauro, text=textoMauro, textinfo='text+percent',
                                        textposition='inside', insidetextfont=dict(size=14, color='white'),
                                        showlegend=False, marker=dict(colors=cores)), row=1, col=1)

    graficoCoordenadores.add_trace(go.Pie(labels=status, values=rogerio, hovertemplate=hoverTemplateRogerio, text=textoRogerio, textinfo='text+percent',
                                        textposition='inside', insidetextfont=dict(size=14, color='white'),
                                        showlegend=False, marker=dict(colors=cores)), row=1, col=2)
    
    graficoCoordenadores.add_trace(go.Pie(labels=status, values=marcos, hovertemplate=hoverTemplateMarcos, text=textoMarcos, textinfo='text+percent', 
                                        textposition='inside', insidetextfont=dict(size=14, color='white'),
                                        showlegend=False, marker=dict(colors=cores)), row=1, col=3)

    graficoCoordenadores.add_trace(go.Pie(labels=status, values=ana, hovertemplate=hoverTemplateAna, text=textoAna, textinfo='text+percent', 
                                        textposition='inside', insidetextfont=dict(size=14, color='white'),
                                        showlegend=False, marker=dict(colors=cores)), row=1, col=4)

    for i, titulo in enumerate(['MAURO', 'ROGÉRIO', 'MARCOS', 'ANA'], start=1):
        graficoCoordenadores.update_xaxes(title_text=titulo, row=1, col=i)
    # Atualizando as anotações e o layout
    graficoCoordenadores.update_annotations(font=dict(family='Gill Sans MT', size=22, color='#2c4594'))
    graficoCoordenadores.update_layout(margin=dict(t=0, b=50, l=0, r=70),
        title={
            'text': 'STATUS DOS PROJETOS POR COORDENADOR',
            'x': 0.05,
            'y': 0.9,
            'font': {
            'family': 'Gill Sans MT',
            'color': '#2c4594',
            'size': 28
            }},
    )
    
    graficoCoordenadores.add_layout_image(
        dict(
            source=bg,
            xref="paper",
            yref="paper",
            x=0,
            y=1,
            sizex=2,
            sizey=2,
            sizing="stretch",
            opacity=0.5,
            layer="below"
        )
    )
    # Setando templates
    graficoCoordenadores.update_layout(template="plotly_white")

    return graficoCoordenadores, graficoProjetos