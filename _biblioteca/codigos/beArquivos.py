# Importando as bibliotecas necessárias
import locale
import pandas as pd
import os 
import smtplib
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime

# Definindo a configuração regional para datas em português
locale.setlocale(locale.LC_TIME, 'pt_PT.UTF-8')

# Função para abrir planilha e ler todas suas abas, retornando em lista de dict
def f_abrePlanilha(caminhoArquivo):
    # Lendo arquivo
    arquivoSelecionado = pd.ExcelFile(caminhoArquivo)
    nomesAbas = arquivoSelecionado.sheet_names
    # Lista para armazenar os dados de todas as abas
    tabelaLida = []
    dictAcumulado = {}
    # Obtendo a data atual como um objeto Timestamp
    hoje = pd.Timestamp.now().normalize()  # Normaliza para começar no início do dia

    # Percorrendo cada aba
    for indiceAbas, tabelaDaVez in enumerate(arquivoSelecionado.sheet_names):
        # Lendo a aba atual
        tabelaDaVez = arquivoSelecionado.parse(
            tabelaDaVez,

            # Pegando apenas colunas A e B
            usecols=[0, 1, 2],

            # Nomeando abas no DataFrame
            names=['tarefas', 'datas', 'conclusao']
        )
        porcentagemConclusao = tabelaDaVez['conclusao'].tolist()
        # Convertendo datas de timestamp para string personalizada
        tabelaDaVez['datas'] = tabelaDaVez['datas'].apply(lambda dataDaVez: dataDaVez if not pd.isnull(dataDaVez) else None)
        # Adicionando colunas de status e cor do status
        nLinhas = len(tabelaDaVez)
        tabelaDaVez.insert(0, 'cor', ['f7f7f7'] * nLinhas)
        # Adicionando coluna de status com valores padrão 'OK'
        tabelaDaVez.insert(1, 'status', ['OK'] * nLinhas)
        tabelaDaVez.insert(4, 'desvios', '') 
        tabelaDaVez.insert(5, 'planoDeAcao', [' '] * nLinhas)
        tabelaDaVez.insert(6, 'dataDoPlano', [' '] * nLinhas)
        tabelaDaVez.insert(7, 'coordenador', [' '] * nLinhas)
        tabelaDaVez.insert(9, 'statusPlano', '')

        # Verificando se a tarefa já está atrasada ou em risco
        for i, data in enumerate(tabelaDaVez['datas']):
            if not pd.isnull(data):
                diferencaDias = (data - hoje).days
                if diferencaDias < 0:
                    tabelaDaVez.at[i, 'status'] = 'ATRASADO'
                elif diferencaDias <= 3:
                    tabelaDaVez.at[i, 'status'] = 'RISCO'
                # Percorrendo a lista de porcentagens de conclusão
            for i, conclusao in enumerate(porcentagemConclusao):
                # Verificando se a conclusão é igual a 100
                if conclusao == 100:
                    # Atualizando o status para 'ENTREGUE'
                    tabelaDaVez.at[i, 'status'] = 'ENTREGUE'
        # Adicionando os dados da aba atual à lista
        tabelaLida.append({
            'gestor': nomesAbas[indiceAbas].upper(), 
            'dados': tabelaDaVez,
            'dictAcumulado': {}
        })

    return tabelaLida

# Função para salvar os DataFrames em arquivos .xlsx
def f_salvaExcel(df, gestorDaVez):
    email = ""
    # identificando gestor
    if gestorDaVez == 'MARCOS':
        arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos/planoMARCOS.xlsx'
        email = "marcos.freitas@vibracon.com.br"
    elif gestorDaVez == 'ROGERIO':
        arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos/planoROGERIO.xlsx'
        email = "rogerio.queiroz@vibracon.com.br"
    elif gestorDaVez == 'MAURO':
        arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos/planoMAURO.xlsx'
        email = "mauro.barbosa@vibracon.com.br"
    else:
        arquivo = 'W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos/planoANA.xlsx'
        email = "ana.flavia@vibracon.com.br"

    # carregando planilha existente 
    wb = load_workbook(arquivo)
    planilha = wb.active
    dadosNovos = df.iloc[:, [-5, -4, -3, -1]].values
    colunaInicio = 2
    larguraColunas = {'A': 55, 'B': 40, 'C': 20, 'D': 20, 'E': 20}

    # Inserir os dados na planilha Excel
    for linha, dadosLinha in enumerate(dadosNovos, start=1):
        for coluna, dado in enumerate(dadosLinha, start=colunaInicio):
            planilha.cell(row=linha+1, column=coluna, value=dado)

    for col, width in larguraColunas.items():
        planilha.column_dimensions[col].width = width

    planilha.cell(row=1, column=2, value='PLANO DE AÇÃO')
    planilha.cell(row=1, column=3, value='DATA')
    planilha.cell(row=1, column=4, value='RESPONSÁVEL')
    planilha.cell(row=1, column=5, value='STATUS DO PLANO')
    planilha.delete_cols(planilha.max_column + 1, planilha.max_column + 5)
    # Salvar o arquivo Excel com as alterações
    wb.save(arquivo)
    destinatarios = ["fernando.teixeira@vibracon.com.br", email]
    #destinatario = "lucassteix1@gmail.com"
    for destinatario in destinatarios:
        f_enviarPlanoDeAcao(destinatario, arquivo)

# ---------------------------------------------------
# Função para enviar um e-mail com o plano de ação atualizado
def f_enviarPlanoDeAcao(destinatario, arquivo):
    
    dataAtual = datetime.today().date().strftime('%d-%m-%Y')
    # Parâmetros para enviar o e-mail
    remetente = "vibraconvibraplan@gmail.com"
    senhaRemetente = "ncha qtxw zdmu ixjt"
    assunto = f"Plano de ação {dataAtual}"  
    mensagem = f"Prezado, segue em anexo o plano de ação do dia {dataAtual}."
    
    # Configurando a mensagem de e-mail
    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = destinatario
    msg['Subject'] = assunto

    msg.attach(MIMEText(mensagem, 'plain'))

    # Anexando o arquivo ao e-mail
    with open(arquivo, 'rb') as f:
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(arquivo)}')
        msg.attach(attachment)

    # Enviando o e-mail
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587) 
        server.starttls()
        server.login(remetente, senhaRemetente)
        text = msg.as_string()
        server.sendmail(remetente, destinatario, text)
        server.quit()
    except Exception as e:
        print(f"Erro ao enviar o e-mail: {str(e)}")    

# ---------------------------------------------------
# Função para enviar um e-mail com o plano de ação mensal
def f_enviarPlanoMensal(destinatario):

    mes = datetime.now().month
    # Parâmetros para enviar o e-mail
    remetente = "vibraconvibraplan@gmail.com"
    senhaRemetente = "ncha qtxw zdmu ixjt"
    arquivoTxt = "W:/PLANEJAMENTO/GESTÃO À VISTA/Vibraplan/_arquivos/Planos/planoMensal.txt"
    assunto = f"Plano de ação {mes}"  
    mensagem = f"Prezado, segue em anexo o plano de ação mensal."
    
    # Configurando a mensagem de e-mail
    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = destinatario
    msg['Subject'] = assunto

    msg.attach(MIMEText(mensagem, 'plain'))

    # Anexando o arquivo ao e-mail
    with open(arquivoTxt, 'rb') as f:
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(arquivoTxt)}')
        msg.attach(attachment)

    # Enviando o e-mail
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587) 
        server.starttls()
        server.login(remetente, senhaRemetente)
        text = msg.as_string()
        server.sendmail(remetente, destinatario, text)
        server.quit()
    except Exception as e:
        print(f"Erro ao enviar o e-mail: {str(e)}")   